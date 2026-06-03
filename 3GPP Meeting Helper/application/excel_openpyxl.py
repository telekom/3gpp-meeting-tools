import zipfile
from typing import NamedTuple, Dict

import openpyxl
import openpyxl.utils
from openpyxl.utils import column_index_from_string
from openpyxl.utils.cell import coordinate_from_string
import pandas as pd
import xml.etree.ElementTree as ET


def parse_excel_hyperlinks_by_column_name(file_path, column_name, sheet_name=None):
    """
    Parses hyperlinks from a specific column in an Excel (.xlsx) file,
    identified by its column name (header).

    Args:
        file_path (str): The path to the Excel file.
        column_name (str): The name of the column (header) to parse.
        sheet_name (str, optional): The name of the sheet to parse.
                                     If None, it will parse the active sheet.

    Returns:
        list: A list of tuples, where each tuple contains (hyperlink_text, hyperlink_target).
              Returns an empty list if no hyperlinks are found or an error occurs.
    """
    hyperlinks_info = set()
    column_letter = None

    try:
        workbook = openpyxl.load_workbook(file_path)

        if sheet_name:
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
            else:
                print(f"Error: Sheet '{sheet_name}' not found in the workbook.")
                return hyperlinks_info
        else:
            sheet = workbook.active

        print(f"Parsing column '{column_name}' in sheet: {sheet.title}")

        # 1. Find the column index for the given column_name
        col_index = None
        for col_idx, cell in enumerate(sheet[1]):
            if cell.value == column_name:
                col_index = col_idx
                break

        if col_index is None:
            print(f"Error: Column '{column_name}' not found in the header row.")
            return hyperlinks_info

        try:
            column_letter = openpyxl.utils.get_column_letter(col_index + 1)
        except ValueError:
            print(f"Error: Column '{column_name}' not found in the header row.")
            return hyperlinks_info

        print(f"Column '{column_name}' corresponds to column letter '{column_letter}'.")

        # 2. Iterate directly over the sheet's hyperlinks collection
        for hl in sheet.hyperlinks:
            if not hl.ref:
                continue

            # hl.ref can sometimes be a merged range like 'C5:C10'.
            # Splitting by ':' and taking the first part guarantees we get the top-left cell ('C5')
            top_left_coord = hl.ref.split(':')[0]

            # Robustly separate the letter and the number (e.g., 'C' and 5)
            hl_col, hl_row = coordinate_from_string(top_left_coord)

            # STRICT FILTER: Only process if it matches our target column AND is not the header row
            if hl_col == column_letter and hl_row > 1:
                # We only instantiate the cell object in Python if it passes the filter
                cell_value = sheet[top_left_coord].value

                if hl.target:
                    # 3. Use .add() instead of .append().
                    # If it's a duplicate, the set silently ignores it in O(1) time.
                    hyperlinks_info.add((cell_value, hl.target))

    except FileNotFoundError:
        print(f"Error: File not found at {file_path}")
    except Exception as e:
        print(f"An error occurred: {e}")

    # Remove duplicates by converting to a set and then back to a list
    unique_hyperlinks = list(hyperlinks_info)

    # Sort the list of tuples alphabetically by the hyperlink's text
    unique_hyperlinks.sort(key=lambda x: str(x[0]) if x[0] is not None else "")

    return unique_hyperlinks


def extract_hyperlinks_zip_merge(excel_path: str, df: pd.DataFrame, column_name: str,
                                 sheet_filename="sheet1.xml") -> list:
    """
    Instantly extracts hyperlinks by unzipping the Excel file and mapping
    the XML relationships directly to the already-loaded Pandas Dataframe.
    """
    if column_name not in df.columns:
        print(f"Error: Column '{column_name}' not found in dataframe.")
        return []

    # 1. Map the Excel column layout mathematically.
    # Since you use `index_col=0` in pd.read_excel:
    # Excel Col A -> df.index
    # Excel Col B -> df.columns[0]
    # Therefore, Excel Column Index = df index + 2
    df_col_idx = df.columns.get_loc(column_name)
    excel_target_col_num = df_col_idx + 2

    hyperlinks_info = set()

    # XML Namespaces used by Excel
    ns = {
        'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
        'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
        'rels': 'http://schemas.openxmlformats.org/package/2006/relationships'
    }

    try:
        # 2. Extract raw XML from the Zip archive directly into memory
        with zipfile.ZipFile(excel_path, 'r') as z:

            # Step A: Get the hyperlink targets from the .rels file
            try:
                rels_xml = z.read(f"xl/worksheets/_rels/{sheet_filename}.rels")
            except KeyError:
                return []  # File contains no hyperlinks

            rels_tree: ET.Element = ET.fromstring(rels_xml)
            rel_map = {
                rel.attrib['Id']: rel.attrib['Target']
                for rel in rels_tree.findall('rels:Relationship', ns)
                if "hyperlink" in rel.attrib.get('Type', '')
            }

            # Step B: Get the cell coordinates from the main sheet file
            sheet_xml = z.read(f"xl/worksheets/{sheet_filename}")
            sheet_tree = ET.fromstring(sheet_xml)

            # 3. Filter and Merge
            for hl in sheet_tree.findall('.//main:hyperlink', ns):
                ref = hl.attrib.get('ref')
                r_id = hl.attrib.get(f"{{{ns['r']}}}id")

                if ref and r_id in rel_map:
                    # Get the top-left cell if it's a merged range
                    top_left_coord = ref.split(':')[0]
                    col_str, row_str = coordinate_from_string(top_left_coord)
                    col_num = column_index_from_string(col_str)
                    row_num = int(row_str)

                    # strictly filter by our target column and skip header row
                    if col_num == excel_target_col_num and row_num > 1:
                        target_url = rel_map[r_id]

                        # Map the Excel row to the Pandas iloc index
                        # Excel Row 2 corresponds to df.iloc[0]
                        try:
                            cell_value = df.iloc[row_num - 2, df_col_idx]

                            # Handle Pandas NaNs for empty cells
                            if pd.isna(cell_value):
                                cell_value = None

                            hyperlinks_info.add((cell_value, target_url))
                        except IndexError:
                            pass  # Handles edge cases where XML refs outlast dataframe bounds

    except Exception as e:
        print(f"Zip extraction failed: {e}")
        return []

    unique_links = list(hyperlinks_info)
    unique_links.sort(key=lambda x: str(x[0]) if x[0] is not None else "")

    return unique_links

class WiData(NamedTuple):
    wi_name:str
    wi_url:str

def parse_tdoc_3gu_list_for_wis(file_path) -> Dict[str,str]:
    hyperlinks_list = parse_excel_hyperlinks_by_column_name(file_path=file_path, column_name='Related WIs')
    return dict(hyperlinks_list)


