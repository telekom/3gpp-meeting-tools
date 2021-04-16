import io
import os
import os.path
import re
import traceback

import openpyxl
import pandas as pd
import win32com.client
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font
from openpyxl.styles import PatternFill

win32c = win32com.client.constants

color_magenta = (234, 10, 142)
color_black = (0, 0, 0)
color_white = (255, 255, 255)

# Also used for conditional formatting
color_green = (0, 97, 0)
color_light_green = (198, 239, 206)
color_dark_red = (156, 0, 6)
color_light_red = (255, 199, 206)
color_dark_grey = (128, 128, 128)
color_light_grey = (217, 217, 217)
color_dark_yellow = (156, 87, 0)
color_light_yellow = (255, 235, 156)

comments_regex = re.compile(r'Comment[s]? [\(]?([\w]+)[\)]?|(.*Session) [cC]omments')
comments_filename_regex = re.compile(r'.*[Cc]omments.*\.xlsx')
comments_summary_column = 'Comments summary'
session_comments_column = 'Session comments'
revision_of_column = 'Revision of'
revised_to_column = 'Revised to'

last_column = 'U'


def get_excel():
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = True
        excel.DisplayAlerts = False
    except:
        excel = None
        traceback.print_exc()
    return excel


def open_excel_document(filename=None, excel=None, sheet_name=None):
    if excel is None:
        excel = get_excel()
    if (filename is None) or (filename == ''):
        wb = get_excel().Workbooks.Add()
    else:
        wb = get_excel().Workbooks.Open(filename)
    if sheet_name is not None:
        select_worksheet(wb, sheet_name)
    return wb


def select_worksheet(wb, name):
    wb.Worksheets(name).Activate()


def set_first_row_as_filter(wb, ws_name=None, already_activated=False):
    try:
        if not already_activated:
            wb.Activate()
        if ws_name is None:
            ws = wb.ActiveSheet
        else:
            ws = wb.Sheets(ws_name)
            ws.Activate()
        ws.Range("1:1").AutoFilter()
        ws.Cells(2, 2).Select()
        get_excel().ActiveWindow.FreezePanes = True
    except:
        traceback.print_exc()


def adjust_tdocs_by_agenda_column_width(wb):
    try:
        wb.Activate()
        ws = wb.ActiveSheet
        ws.Range("A:A").ColumnWidth = 11
        ws.Range("B:B").ColumnWidth = 7
        ws.Range("C:C").ColumnWidth = 11
        ws.Range("D:D").ColumnWidth = 12
        ws.Range("E:E").ColumnWidth = 50
        ws.Range("F:F").ColumnWidth = 45
        ws.Range("G:G").ColumnWidth = 8
        ws.Range("H:H").ColumnWidth = 25
        ws.Range("I:I").ColumnWidth = 50
        ws.Range("J:J").ColumnWidth = 11
        ws.Range("K:J").ColumnWidth = 14
        ws.Range("L:L").ColumnWidth = 14
        ws.Range("M:M").ColumnWidth = 14
        ws.Range("N:N").ColumnWidth = 14
        ws.Range("O:O").ColumnWidth = 6

        # CR info
        ws.Range("P:P").ColumnWidth = 7
        ws.Range("Q:Q").ColumnWidth = 7

        # Original and final TDocs
        ws.Range("R:R").ColumnWidth = 30
        ws.Range("S:S").ColumnWidth = 30

        # Source (Summary)
        ws.Range("T:T").ColumnWidth = 45

        # Session comments
        ws.Range("U:U").ColumnWidth = 75

        # General formatting
        ws.Range("A:" + last_column).WrapText = True
        ws.Range("A:" + last_column).EntireRow.AutoFit()
    except:
        traceback.print_exc()


def close_wb(wb):
    wb.Close()


def add_pivot_table_to_workbook_active_sheet(
        wb,
        cell_range,
        data_field=None,
        row_field=None,
        column_field=None,
        ws_name='',
        order_by=''):
    print("Generating pivot table")
    ws = wb.ActiveSheet

    # Generate Pivot chart in new worksheet based on columns A:D
    pivot_table_source_data = ws.Range(cell_range)
    XlPivotTableSourceType = 1  # xlDatabase, see https://docs.microsoft.com/en-us/office/vba/api/excel.xlpivottablesourcetype
    XlPivotTableVersion = 5  # xlPivotTableVersion15 (Excel 2013), see https://docs.microsoft.com/en-us/office/vba/api/excel.xlpivottableversionlist
    pivot_cache = wb.PivotCaches().Create(XlPivotTableSourceType, pivot_table_source_data, XlPivotTableVersion)

    # Create pivot chart in the end (looks better)
    last_wb = wb.Sheets(wb.Sheets.Count)
    pivot_chart_wb = wb.Sheets.Add(Before=None, After=last_wb)
    pivot_chart_wb.Name = ws_name

    pivot_table = pivot_cache.CreatePivotTable(
        TableDestination=pivot_chart_wb.Range("A1"),
        TableName='{0} Pivot Table'.format(ws_name))

    # Copied from https://stackoverflow.com/questions/46312435/create-a-pivotchart-with-python-win32com-causes-pywintypes-com-error
    pivot_table.ColumnGrand = True
    pivot_table.HasAutoFormat = True
    pivot_table.DisplayErrorString = False
    pivot_table.DisplayNullString = True
    pivot_table.EnableDrilldown = True
    pivot_table.ErrorString = ""
    pivot_table.MergeLabels = False
    pivot_table.NullString = ""
    pivot_table.PageFieldOrder = 2
    pivot_table.PageFieldWrapCount = 0
    pivot_table.PreserveFormatting = True
    pivot_table.RowGrand = True
    pivot_table.SaveData = True
    pivot_table.PrintTitles = False
    pivot_table.RepeatItemsOnEachPrintedPage = True
    pivot_table.TotalsAnnotation = False
    pivot_table.CompactRowIndent = 1
    pivot_table.InGridDropZones = False
    pivot_table.DisplayFieldCaptions = True
    pivot_table.DisplayMemberPropertyTooltips = False
    pivot_table.DisplayContextTooltips = True
    pivot_table.ShowDrillIndicators = True
    pivot_table.PrintDrillIndicators = False
    pivot_table.AllowMultipleFilters = False
    pivot_table.SortUsingCustomLists = True
    pivot_table.FieldListSortAscending = False
    pivot_table.ShowValuesRow = False
    pivot_table.CalculatedMembersInFilters = False

    pivot_fields = {}

    # See https://docs.microsoft.com/en-us/office/vba/api/excel.xlpivotfieldorientation
    # xlColumnField	2	Column
    # xlDataField	4	Data
    # xlHidden	0	Hidden
    # xlPageField	3	Page
    # xlRowField	1	Row
    if data_field is not None:
        data_field_pivot = pivot_table.PivotFields(data_field)
        data_field_pivot.Orientation = 4
        data_field_pivot.Position = 1
        pivot_fields[data_field] = data_field_pivot

    if row_field is not None:
        row_field_pivot = pivot_table.PivotFields(row_field)
        row_field_pivot.Orientation = 1
        row_field_pivot.Position = 1
        pivot_fields[row_field] = row_field_pivot

    if column_field is not None:
        column_field_pivot = pivot_table.PivotFields(column_field)
        column_field_pivot.Orientation = 2
        column_field_pivot.Position = 1
        pivot_fields[column_field] = column_field_pivot

    if order_by in pivot_fields:
        # See:
        #   - https://docs.microsoft.com/en-us/office/vba/api/excel.xlsortorder
        #   - https://docs.microsoft.com/en-us/office/vba/api/excel.pivotfield.autosort
        # xlAscending = 1
        xlDescending = 2
        # xlManual = -4135
        pivot_fields[order_by].AutoSort(xlDescending)


def generate_pivot_chart_from_tdocs_by_agenda(wb):
    try:
        wb.Activate()
        ws = wb.ActiveSheet
        add_pivot_table_to_workbook_active_sheet(
            wb,
            cell_range="A:D",
            data_field="TD#",
            row_field="AI",
            column_field="Type",
            ws_name="AI Summary",
            order_by='AI')

        print("Generating per-AI pivot chart")
        points_per_cm = 28.3465
        chart_width_cm = 24
        chart_height_cm = 14
        summary_chart = wb.Charts.Add()
        # summary_chart = summary_shape.Chart
        summary_chart.Name = "TDocs per AI"
        summary_chart.ChartType = 52  # xlColumnStacked, see https://docs.microsoft.com/en-us/office/vba/api/excel.xlcharttype
        summary_chart.ChartArea.Format.Line.Visible = 0  # True = -1; False = 0
        summary_chart.HasTitle = False
        summary_chart.HasLegend = True  # Legend show because we are showing per-type stacked bars
        summary_chart.ShowAllFieldButtons = False
        x_axis = summary_chart.Axes(1)
        x_axis.HasTitle = True
        x_axis.AxisTitle.Text = "AI"
        y_axis = summary_chart.Axes(2)
        y_axis.HasTitle = True
        y_axis.AxisTitle.Text = "TDocs"

        # Set the original Worksheet as active
        ws.Activate()
    except:
        traceback.print_exc()


def vertically_center_all_text(wb):
    try:
        wb.Activate()
        ws = wb.ActiveSheet
        # Constants do not work well with win32com, so we just use the value directly
        # https://docs.microsoft.com/en-us/office/vba/api/excel.xlvalign
        ws.Range("A:" + last_column).EntireRow.VerticalAlignment = -4108
    except:
        traceback.print_exc()


# https://stackoverflow.com/questions/11444207/setting-a-cells-fill-rgb-color-with-pywin32-in-excel
def rgb_to_hex(rgb):
    '''
    ws.Cells(1, i).Interior.color uses bgr in hex

    '''
    bgr = (rgb[2], rgb[1], rgb[0])
    strValue = '%02x%02x%02x' % bgr
    # print(strValue)
    iValue = int(strValue, 16)
    return iValue


def hide_columns(wb, columns):
    try:
        wb.Activate()
        ws = wb.ActiveSheet

        for column in columns:
            print('Hiding column {0}'.format(column))
            ws.Columns(column).Hidden = True
    except:
        traceback.print_exc()


def set_tdoc_colors(wb, links, no_index_links=True):
    tdoc_cell_mapping = {}
    try:
        wb.Activate()
        ws = wb.ActiveSheet

        #  Get last row
        last_row = ws.Rows.Count

        # Borders
        ws.Range("A1:" + last_column + str(
            last_row)).Borders.LineStyle = 1  # xlContinuous -> https://docs.microsoft.com/en-us/office/vba/api/excel.xllinestyle
        ws.Range("A1:" + last_column + str(last_row)).Borders.Color = rgb_to_hex(color_black)
        ws.Range("A1:" + last_column + str(
            last_row)).Borders.Weight = 2  # xlThin -> https://docs.microsoft.com/en-us/office/vba/api/excel.xlborderweight

        # Results conditional formatting
        # https://docs.microsoft.com/en-us/office/vba/api/excel.xlformatconditiontype
        # xlCellValue 1
        # xlEqual 3
        results_cells = "J2:J" + str(last_row)

        def apply_conditional_formatting(the_range, criteria, text_color, background_color):
            the_cells = ws.Range(the_range).FormatConditions.Add(1, 3, criteria)
            the_cells.Font.Color = rgb_to_hex(text_color)
            the_cells.Interior.Color = rgb_to_hex(background_color)
            return the_cells

        # Results formatting
        apply_conditional_formatting(results_cells, 'Agreed', color_green, color_light_green)
        apply_conditional_formatting(results_cells, 'Approved', color_green, color_light_green)
        apply_conditional_formatting(results_cells, 'Noted', color_dark_red, color_light_red)
        apply_conditional_formatting(results_cells, 'Revised', color_dark_grey, color_light_grey)
        apply_conditional_formatting(results_cells, 'Merged', color_dark_grey, color_light_grey)
        apply_conditional_formatting(results_cells, 'Postponed', color_dark_yellow, color_light_yellow)
    except:
        traceback.print_exc()


def save_wb(wb):
    try:
        wb.Activate()
        wb.Save()
        print('Workbook saved!')
    except:
        traceback.print_exc()


def get_company_name_based_on_email(sender_address):
    company_name = ''
    try:
        split_company_name = sender_address.split('@')[-1].split('.')
        company_name = split_company_name[-2].title()
        # Fix for ZTE company name
        if company_name == 'Com' and len(split_company_name) > 2:
            company_name = split_company_name[-3].title()

        # Some capitalization of short company names
        if len(company_name) < 5:
            company_name = company_name.upper()
    except:
        company_name = 'Could not parse'

    return company_name


def export_email_approval_list(local_filename, found_attachments, tdocs_without_emails=None, tdoc_data=None):
    if (local_filename is None) or (local_filename == ''):
        return

    # found_attachments -> collections.namedtuple('RevisionDoc', 'time tdoc filename absolute_url sender_name sender_address chairman_notes')

    print('Starting email approval export: {0} emails'.format(len(found_attachments)))

    # Faster variant writing first most data not using VBA
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet()
    ws.title = "Revisions"

    # Add title row
    ws.append(['TD#', 'Time', 'Filename mention', 'Sender', 'Company', 'Email', 'AI', "Chairman's notes"])

    # Add email entries
    for idx, item in enumerate(found_attachments, start=2):
        filename_cell = WriteOnlyCell(ws, value=item.filename)
        sender_name_cell = WriteOnlyCell(ws, value=item.sender_name)
        link_cell = WriteOnlyCell(ws, value='Link')

        # Link to file. May not always be a path
        if item.absolute_url != '':
            filename_cell.hyperlink = 'file:///' + item.absolute_url
            filename_cell.font = Font(underline="single", color='00EA0A8E')
        # Link to author
        sender_name_cell.hyperlink = 'mailto:' + item.sender_address
        sender_name_cell.font = Font(underline="single", color='00EA0A8E')
        # Link to email
        link_cell.hyperlink = 'file:///' + item.email_url
        link_cell.font = Font(underline="single", color='00EA0A8E')

        # Write row
        ws.append([
            item.tdoc,
            item.time,
            filename_cell,
            sender_name_cell,
            get_company_name_based_on_email(item.sender_address),
            link_cell,
            str(item.ai_folder),
            str(item.chairman_notes)
        ])

    if tdocs_without_emails is not None and tdoc_data is not None:
        ws = wb.create_sheet()
        ws.title = "TDocs without emails"
        ws.append(['TD#', 'AI', 'Type', 'Doc For', 'Title', 'Source', 'Rel', 'Work Item', 'Comments'])

        tdocs_info = tdoc_data.tdocs.loc[
            list(tdocs_without_emails), ['AI', 'Type', 'Doc For', 'Title', 'Source', 'Rel', 'Work Item', 'Comments']]
        print('{0} TDocs without matching emails'.format(len(tdocs_info.index)))
        for row_index, row in tdocs_info.iterrows():
            ws.append([
                row_index,
                str(row['AI']),
                str(row['Type']),
                str(row['Doc For']),
                str(row['Title']),
                str(row['Source']),
                str(row['Rel']),
                str(row['Work Item']),
                str(row['Comments'])])

    print('Saving Excel table structure')
    wb.save(filename=local_filename)
    print('Closing Excel File')
    wb.close()
    print('Excel File closed')

    # Only necessary things with VBA (much slower)
    try:
        print('Applying Excel formatting')
        wb = open_excel_document(filename=local_filename)
        # ws = wb.ActiveSheet
        ws = wb.Sheets("Revisions")

        ws.Range("A:A").ColumnWidth = 14
        ws.Range("B:B").ColumnWidth = 18
        ws.Range("C:C").ColumnWidth = 30  # Reduced to 30 as we now just send around revision numbers
        ws.Range("D:D").ColumnWidth = 40
        ws.Range("E:E").ColumnWidth = 17
        ws.Range("F:F").ColumnWidth = 9
        ws.Range("G:G").ColumnWidth = 25
        ws.Range("H:H").ColumnWidth = 60

        ws.Range("A:B").HorizontalAlignment = -4108
        ws.Range("G:G").HorizontalAlignment = -4108
        ws.Range("A:H").VerticalAlignment = -4108
        ws.Range("A:H").WrapText = True
        ws.Range("H:H").WrapText = True
        ws.Range("A1:H1").Font.Bold = True

        set_first_row_as_filter(wb, 'Revisions')

        ws.AutoFilter.Sort.SortFields.Clear()

        # https://docs.microsoft.com/en-us/office/vba/api/excel.xlsortorder
        xlAscending = 1
        # https://stackoverflow.com/questions/11766118/excel-constants-for-sorting
        xlSortOnValues = 0

        ws.AutoFilter.Sort.SortFields.Add(Order=xlAscending, SortOn=xlSortOnValues, Key=ws.Range("G:G"))  # AI
        ws.AutoFilter.Sort.SortFields.Add(Order=xlAscending, SortOn=xlSortOnValues, Key=ws.Range("A:A"))  # TD
        ws.AutoFilter.Sort.SortFields.Add(Order=xlAscending, SortOn=xlSortOnValues, Key=ws.Range("B:B"))  # Time
        ws.AutoFilter.Sort.Apply()

        if tdocs_without_emails is not None and tdoc_data is not None:
            ws = wb.Sheets("TDocs without emails")

            # TD, AI, Type, Doc For, Title, source, rel, work item, comments
            ws.Range("A:A").ColumnWidth = 11
            ws.Range("B:B").ColumnWidth = 7
            ws.Range("C:C").ColumnWidth = 11
            ws.Range("D:D").ColumnWidth = 12
            ws.Range("E:E").ColumnWidth = 50
            ws.Range("F:F").ColumnWidth = 45
            ws.Range("G:G").ColumnWidth = 7
            ws.Range("H:H").ColumnWidth = 7
            ws.Range("I:I").ColumnWidth = 50

            ws.Range("A:I").HorizontalAlignment = -4108
            ws.Range("A:I").VerticalAlignment = -4108
            ws.Range("A:I").WrapText = True
            ws.Range("A1:I1").Font.Bold = True

            set_first_row_as_filter(wb, 'TDocs without emails', already_activated=True)
            ws.AutoFilter.Sort.SortFields.Clear()
            ws.AutoFilter.Sort.SortFields.Add(Order=xlAscending, SortOn=xlSortOnValues, Key=ws.Range("B:B"))  # AI
            ws.AutoFilter.Sort.SortFields.Add(Order=xlAscending, SortOn=xlSortOnValues, Key=ws.Range("A:A"))  # TD
            ws.AutoFilter.Sort.Apply()

        print('Finished email approval export')

        revisions_ws = wb.Worksheets('Revisions').Activate()
        add_pivot_table_to_workbook_active_sheet(
            wb,
            cell_range="A:G",
            data_field="TD#",
            row_field="Sender",
            column_field=None,
            ws_name="Delegate Summary")
        revisions_ws = wb.Worksheets('Revisions').Activate()
        add_pivot_table_to_workbook_active_sheet(
            wb,
            cell_range="A:G",
            data_field="TD#",
            row_field="Company",
            column_field=None,
            ws_name="Company Summary")
        wb.SaveAs(local_filename)
    except:
        traceback.print_exc()


def read_comments_file(filename):
    try:
        xl = pd.ExcelFile(filename)
        ws = xl.sheet_names[0]
        df = xl.parse(ws, index_col=0)
        column_names = list(df.columns.values)
        if session_comments_column not in column_names:
            df[session_comments_column] = ''
        # Avoid type errors when RegEx-ing
        column_names = [str(e) for e in column_names]
        column_name_matches = [comments_regex.match(column_name) for column_name in column_names]
        cleaned_column_name_matches = [match for match in column_name_matches if match is not None]
        column_matches = [(match_data.group(0), match_data.group(1)) for match_data in cleaned_column_name_matches]
        name_columns = [col_data[1] for col_data in column_matches]
        for idx in df.index.values:
            try:
                initial_comments = []
                for col in column_matches:
                    comment = df.at[idx, col[0]]
                    if (comment is not None) and isinstance(comment, str) and (comment != ''):
                        comment_source = col[1]
                        # None in the case of Session Comments
                        if comment_source is not None:
                            initial_comments.append('[{0}]: {1}'.format(comment_source, comment.strip('\n')))
                row_full_comments = '\n'.join(initial_comments)
                session_comments = str(df.at[idx, session_comments_column])
                if session_comments == 'nan':
                    session_comments = None
                if (session_comments is None) or (session_comments == ''):
                    summary_comments = row_full_comments
                else:
                    summary_comments = '\n\n'.join(
                        [item for item in [row_full_comments, session_comments] if (item is not None) and (item != '')])
                df.at[idx, comments_summary_column] = summary_comments
            except:
                print('Could not import comments for TDoc {1} file {0}'.format(filename, idx))
                traceback.print_exc()
        # Filter out columns with no comments
        df = df.loc[(df[comments_summary_column] is not None) & (df[comments_summary_column] != '')]
        df = df.loc[:, comments_summary_column]
        return df
    except:
        print('Could not import comments file {0}'.format(filename))
        traceback.print_exc()
        return None


def get_comment_data_from_cell(contributor_name, row, idx):
    if contributor_name is not None:
        contributor_name = contributor_name.strip()
    cell = row[idx]
    if cell is None:
        return (contributor_name, None, None, None)
    else:
        # Foreground color
        try:
            fg_color = cell.fill.fgColor.index
        except:
            fg_color = '00000000'
        # Text color
        try:
            font_color = cell.font.color.index
        except:
            font_color = '00000000'

        # ToDo: calculate closeness to red, yellow and green and set the color accordingly
        return (contributor_name, cell.value, fg_color, font_color)


def read_comments_format(filename):
    try:
        # As per https://stackoverflow.com/questions/31416842/openpyxl-does-not-close-excel-workbook-in-read-only-mode
        with open(filename, "rb") as f:
            in_mem_file = io.BytesIO(f.read())

        book = openpyxl.load_workbook(in_mem_file, read_only=True)
        print('Loaded comments file')
        ws = book.active
        comments_row = ws[1]

        column_names = [str(e.value) for e in comments_row]
        column_name_matches = [comments_regex.match(column_name) for column_name in column_names]

        comments_to_map = []
        for idx, match in enumerate(column_name_matches):
            if match is None:
                continue
            comments_to_map.append((match.group(1), idx))

        print('Scanning for comments')
        all_comments = {}
        for row in ws.iter_rows(min_row=2):
            tdoc = row[0].value
            # Fix: cell may be "None"!!!! Replace with loop
            comments = [get_comment_data_from_cell(comment[0], row, comment[1]) for comment in comments_to_map]
            comments = [comment for comment in comments if (comment[1] is not None) and (comment[1] != '')]
            if len(comments) > 0:
                all_comments[tdoc] = comments
        book.close()
        return all_comments
    except:
        print('Could not import comments file {0}'.format(filename))
        traceback.print_exc()
        return None


def get_comments_files_in_dir(directory):
    try:
        all_files = os.listdir(directory)
        comments_files = [filename for filename in all_files if
                          (comments_filename_regex.match(filename) is not None) and (not filename.startswith('~$'))]
        return comments_files
    except:
        return []


def get_comments_from_dir(directory, merge_comments=False):
    files = get_comments_files_in_dir(directory)
    if len(files) == 0:
        return None
    full_df = None
    for file in files:
        print('Importing comments from {0}'.format(file))
        df = read_comments_file(os.path.join(directory, file))
        if full_df is None:
            full_df = df
        else:
            for i, row in df.iteritems():
                if (i in full_df.index) and merge_comments:
                    full_df[i] = full_df[i] + '\n\n' + row
                else:
                    full_df[i] = row
    return full_df


def get_comment_full_text(name, comment):
    if comment is None:
        return ''
    if comment == '':
        return ''
    # Avoid having "[None]" tags 
    if (name is None) or ('Session' in name):
        return '{0}'.format(comment)
    return '[{0}]: {1}'.format(name, comment)


def get_comments_from_dir_format(directory, merge_comments=False):
    files = get_comments_files_in_dir(directory)
    if len(files) == 0:
        return None
    full_comments = None
    for file in files:
        print('Importing comments from {0}'.format(file))
        comments = read_comments_format(os.path.join(directory, file))
        print('Read comments for {0} TDocs'.format(len(comments)))
        if full_comments is None:
            full_comments = comments
        else:
            for tdoc, tdoc_comments in comments.items():
                if tdoc not in full_comments:
                    full_comments[tdoc] = tdoc_comments
                else:
                    # Add only entries not already here
                    existing_comments = full_comments[tdoc]
                    existing_texts = [get_comment_full_text(comment_data[0], comment_data[1]) for comment_data in
                                      existing_comments]
                    for comment_to_eval in tdoc_comments:
                        text_to_eval = get_comment_full_text(comment_to_eval[0], comment_to_eval[1])
                        if text_to_eval not in existing_texts:
                            # Only append comment if it is not there already
                            existing_comments.append(comment_to_eval)
    return full_comments


def get_reddest_color(colors):
    try:
        sorted_colors = sorted(colors, key=lambda x: int(x[2:4], 16))
        return sorted_colors[-1]
    except:
        return '00000000'


def get_colors_from_comments(comments):
    fg_colors = {}
    text_colors = {}
    if comments is None:
        return fg_colors, text_colors
    for tdoc, comments in comments.items():
        fg_colors_comments = [comment_data[2] for comment_data in comments]
        text_colors_comments = [comment_data[3] for comment_data in comments]

        reddest_fg_color = get_reddest_color(fg_colors_comments)
        reddest_text_color = get_reddest_color(text_colors_comments)

        fg_colors[tdoc] = reddest_fg_color
        text_colors[tdoc] = reddest_text_color
    return fg_colors, text_colors


def apply_comments_coloring_and_hyperlinks(filename, fg_colors, text_colors, server_urls, hyperlink_columns=[revision_of_column, revised_to_column]):
    book = openpyxl.load_workbook(filename)
    ws = book.active
    print('Applying comment color formatting')
    if server_urls is None:
        server_urls = {}
    else:
        try:
            server_urls = dict(server_urls)
        except:
            server_urls = {}
            print('Could not generate TDoc URL mapping')
            traceback.print_exc()
    # Column 20 is the one with the comments
    header_row = [cell.value for cell in ws[1]]
    try:
        session_comments_idx = header_row.index(session_comments_column)
        comments_present = True
    except:
        comments_present = False
    revision_of_idx = header_row.index(revision_of_column)
    revised_to_idx = header_row.index(revised_to_column)

    hyperlink_columns = [e for e in hyperlink_columns if e in header_row]
    hyperlink_column_idxs = [header_row.index(e) for e in header_row]

    for row in ws.iter_rows(min_row=2):
        tdoc = row[0].value
        set_tdoc_hyperlink(tdoc, row[0], server_urls)

        # Linking of revision documents
        for idx in hyperlink_column_idxs:
            cell_to_format = row[idx]
            set_tdoc_hyperlink(cell_to_format.value, cell_to_format, server_urls)

        if tdoc not in fg_colors:
            continue

        # Format comments
        if comments_present:
            set_comments_color(tdoc, row[session_comments_idx], fg_colors, text_colors)
    book.save(filename)


def set_comments_color(tdoc, cell, fg_colors, text_colors):
    try:
        comment_color = fg_colors[tdoc]
        text_color = text_colors[tdoc]
        if (comment_color != '00000000') and (comment_color != 'FFFFFFFF'):
            cell.fill = PatternFill(start_color=comment_color, end_color=comment_color, fill_type='solid')
        if (text_color != '00000000') and (text_color != 'FFFFFFFF'):
            cell.font = Font(color=text_color)
    except:
        print('Could not set color for TDoc {0}'.format(tdoc))
        traceback.print_exc()


def set_tdoc_hyperlink(cell_content, cell, cell_content_to_url_mapping):
    if (cell_content is None) or (cell_content == ''):
        return
    if cell_content in cell_content_to_url_mapping:
        cell.hyperlink = cell_content_to_url_mapping[cell_content]
        cell.font = Font(color='FFEA0A8E', underline='single')

def generate_emeeting_link_for_tdoc(target_cell):
    pass