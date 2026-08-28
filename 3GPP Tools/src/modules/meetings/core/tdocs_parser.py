# --- File: src/modules/meetings/core/tdocs_parser.py ---
import io
import re
import openpyxl
import logging
import json
import os
from pathlib import Path
import docx


class TDocsParser:
    @staticmethod
    def parse_tdocs_excel(filepath: str) -> list:
        json_cache = filepath + ".json"
        try:
            if os.path.exists(json_cache) and os.path.getmtime(json_cache) >= os.path.getmtime(filepath):
                with open(json_cache, "r", encoding="utf-8") as f:
                    return json.load(f)
        except Exception as e:
            logging.warning(f"Could not read JSON cache: {e}")

        data = []
        try:
            with open(filepath, "rb") as f:
                in_mem_file = io.BytesIO(f.read())

            wb = openpyxl.load_workbook(in_mem_file, data_only=True, read_only=True)
            sheet = wb["TDoc_List"] if "TDoc_List" in wb.sheetnames else wb.worksheets[0]

            headers = []
            header_row_idx = 1

            for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
                row_strs = [str(c).strip() if c is not None else "" for c in row]

                hits = 0
                for c in row_strs:
                    cu = c.upper()
                    if cu in ["TDOC", "TD#", "TDOC#"]: hits += 1
                    if cu == "TITLE": hits += 1
                    if cu == "SOURCE": hits += 1
                    if cu == "TYPE": hits += 1
                    if cu == "FOR": hits += 1
                    if "AGENDA ITEM" in cu or cu in ["AI", "AI#", "AI #"]: hits += 1
                    if "STATUS" in cu: hits += 1

                if hits >= 3:
                    header_row_idx = row_idx
                    for val in row_strs:
                        val_clean = re.sub(r'\s+', ' ', val).strip()
                        val_up = val_clean.upper()

                        if ("AGENDA ITEM" in val_up or val_up in ["AI", "AI#",
                                                                  "AI #"]) and "SORT" not in val_up and "DESCRIPTION" not in val_up:
                            val = "Agenda Item"
                        elif val_up in ["TD#", "TDOC#", "TDOC"]:
                            val = "TDoc"
                        # ---> THE FIX: Force the Excel column to map to "TDoc Status"
                        elif "STATUS" in val_up or "RESULT" in val_up:
                            val = "TDoc Status"

                        headers.append(val)
                    break

            if not headers:
                logging.warning("Could not find a valid header row in the TDocs Excel file.")
                wb.close()
                return []

            for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
                row_dict = {}
                is_empty = True
                for i, value in enumerate(row):
                    if i < len(headers) and headers[i]:
                        val_str = str(value).strip() if value is not None else ""
                        row_dict[headers[i]] = val_str
                        if val_str: is_empty = False

                if not is_empty:
                    data.append(row_dict)

            wb.close()

            try:
                with open(json_cache, "w", encoding="utf-8") as f:
                    json.dump(data, f, indent=2)
            except Exception as e:
                logging.warning(f"Could not save JSON cache: {e}")

            return data

        except Exception as e:
            logging.error(f"Failed to parse Excel file {filepath}: {e}")
            return []

    @classmethod
    def parse_tdocs_by_agenda(cls, filepath: str, ui_logger=None) -> dict:
        from bs4 import BeautifulSoup
        import logging

        if ui_logger: ui_logger.emit("⏳ Parsing TdocsByAgenda HTML (Word Export)...", logging.INFO)
        data = {}

        try:
            with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
                soup = BeautifulSoup(f, 'html.parser')

            tables = soup.find_all('table')
            if not tables:
                if ui_logger: ui_logger.emit("❌ No tables found in HTML.", logging.ERROR)
                return data

            target_table = None
            for table in tables:
                headers = [th.get_text(strip=True).lower() for th in table.find_all(['th', 'td'])[:20]]
                if any('td#' in h or 'td #' in h for h in headers):
                    target_table = table
                    break

            if not target_table:
                if ui_logger: ui_logger.emit("❌ Could not identify the main TDoc table in HTML.", logging.ERROR)
                return data

            rows = target_table.find_all('tr')
            if not rows: return data

            header_row = rows[0].find_all(['th', 'td'])
            headers = [h.get_text(separator=' ', strip=True).lower() for h in header_row]

            td_idx = next((i for i, h in enumerate(headers) if 'td#' in h or 'td #' in h), -1)
            comments_idx = next((i for i, h in enumerate(headers) if 'comments' in h), -1)
            email_idx = next((i for i, h in enumerate(headers) if 'e-mail_discussion' in h), -1)
            source_idx = next((i for i, h in enumerate(headers) if 'source' in h), -1)

            # ---> THE FIX: Smarter hunting for HTML columns
            title_idx = next((i for i, h in enumerate(headers) if 'title' in h or 'subject' in h), -1)
            for_idx = next((i for i, h in enumerate(headers) if h == 'for' or 'doc for' in h), -1)
            result_idx = next((i for i, h in enumerate(headers) if 'result' in h or 'status' in h), -1)
            type_idx = next((i for i, h in enumerate(headers) if 'type' in h), -1)

            if td_idx == -1:
                if ui_logger: ui_logger.emit("❌ 'TD#' column missing in HTML table.", logging.ERROR)
                return data

            for row in rows[1:]:
                cols = row.find_all(['td'])
                if len(cols) <= td_idx: continue

                tdoc_id = cols[td_idx].get_text(separator=' ', strip=True)
                if not tdoc_id or not tdoc_id.startswith(('S2-', 'R', 'C', 'S')):
                    continue

                comments = cols[comments_idx].get_text(separator='\n', strip=True) if comments_idx != -1 and len(
                    cols) > comments_idx else ""
                email_disc = cols[email_idx].get_text(separator='\n', strip=True) if email_idx != -1 and len(
                    cols) > email_idx else ""
                title = cols[title_idx].get_text(separator='\n', strip=True) if title_idx != -1 and len(
                    cols) > title_idx else ""
                source = cols[source_idx].get_text(separator='\n', strip=True) if source_idx != -1 and len(
                    cols) > source_idx else ""
                doc_for = cols[for_idx].get_text(separator='\n', strip=True) if for_idx != -1 and len(
                    cols) > for_idx else ""
                result = cols[result_idx].get_text(separator='\n', strip=True) if result_idx != -1 and len(
                    cols) > result_idx else ""
                doc_type = cols[type_idx].get_text(separator='\n', strip=True) if type_idx != -1 and len(
                    cols) > type_idx else ""

                if ui_logger and (comments or email_disc):
                    ui_logger.emit(f"   ➔ Extracted agenda remarks for {tdoc_id}", logging.DEBUG)

                data[tdoc_id] = {
                    'Comments': comments,
                    'e-mail_Discussion': email_disc,
                    'Title': title,
                    'Source': source,
                    'For': doc_for,
                    'Result': result,
                    'Type': doc_type
                }

            if ui_logger: ui_logger.emit(f"✅ Successfully parsed {len(data)} TDocs from Agenda HTML.", logging.INFO)

        except Exception as e:
            if ui_logger: ui_logger.emit(f"❌ Error parsing HTML: {str(e)}", logging.ERROR)

        return data

    @staticmethod
    def parse_tdocs_from_docx(docx_path: str) -> dict:
        """
        Parses 3GPP SA2 document lists (TdocsByAgenda / Chair's Notes) from a .docx table.
        Returns a dictionary compatible with TDocsTableModel.merge_agenda_data().
        """
        path = Path(docx_path)
        if not path.exists():
            logging.error(f"[TDocsParser] File not found: {docx_path}")
            return {}

        try:
            doc = docx.Document(str(path))
        except Exception as e:
            logging.error(f"[TDocsParser] Failed to open docx file '{docx_path}': {e}")
            return {}

        agenda_data = {}

        for table in doc.tables:
            if len(table.rows) < 2:
                continue

            # Identify header row and map column indexes
            header_cells = [cell.text.strip().lower() for cell in table.rows[0].cells]
            col_map = {}

            for idx, text in enumerate(header_cells):
                clean_text = text.replace("\n", " ").strip()
                if clean_text in ['td#', 'td #', 'tdoc', 'td no', 'td no.', 'temporary document']:
                    col_map['tdoc'] = idx
                elif clean_text in ['subject', 'title']:
                    col_map['title'] = idx
                elif clean_text in ['source']:
                    col_map['source'] = idx
                elif clean_text in ['type']:
                    col_map['type'] = idx
                elif clean_text in ['doc for', 'for']:
                    col_map['for'] = idx
                elif clean_text in ['comments', 'comment', 'secretary remarks']:
                    col_map['comments'] = idx
                elif clean_text in ['result', 'status']:
                    col_map['result'] = idx
                elif clean_text in ['ai', 'agenda item']:
                    col_map['ai'] = idx

            # Must have at least a TD column to be a valid TDocs table
            if 'tdoc' not in col_map:
                continue

            current_ai = 'N/A'

            for row in table.rows[1:]:
                cells = [c.text.strip() for c in row.cells]
                if not cells or len(cells) <= col_map['tdoc']:
                    continue

                tdoc_id = cells[col_map['tdoc']].strip()

                # Track current Agenda Item across multi-row sub-headers
                if 'ai' in col_map and col_map['ai'] < len(cells):
                    ai_val = cells[col_map['ai']].strip()
                    if ai_val and ai_val != '-':
                        current_ai = ai_val

                # Skip header repetitions, non-TDoc placeholder rows, and separators
                if not tdoc_id or tdoc_id == '-' or not re.match(r'^[A-Za-z0-9]+-\d+', tdoc_id):
                    continue

                tdoc_id = tdoc_id.upper()
                title = cells[col_map['title']] if 'title' in col_map and col_map['title'] < len(cells) else ''
                source = cells[col_map['source']] if 'source' in col_map and col_map['source'] < len(cells) else ''
                doc_type = cells[col_map['type']] if 'type' in col_map and col_map['type'] < len(cells) else ''
                doc_for = cells[col_map['for']] if 'for' in col_map and col_map['for'] < len(cells) else ''
                comments = cells[col_map['comments']] if 'comments' in col_map and col_map['comments'] < len(
                    cells) else ''
                result = cells[col_map['result']] if 'result' in col_map and col_map['result'] < len(cells) else ''

                agenda_data[tdoc_id] = {
                    'Title': title,
                    'Source': source,
                    'Type': doc_type,
                    'For': doc_for,
                    'Comments': comments,
                    'Result': result,
                    'Agenda Item': current_ai,
                    'e-mail_Discussion': ''
                }

        logging.info(f"[TDocsParser] Successfully parsed {len(agenda_data)} TDocs from {path.name}")
        return agenda_data