# --- File: src/modules/meetings/core/tdocs_merger.py ---
import os
import re
from pathlib import Path
import pandas as pd
from PyQt5.QtCore import QThread, pyqtSignal
import logging

import openpyxl
from openpyxl.styles import Font, Alignment, GradientFill
from openpyxl.styles.fills import Stop
from openpyxl.utils import get_column_letter

from core.network.session import NetworkSession
from modules.meetings.core.tdocs_parser import TDocsParser


class TDocsMergerThread(QThread):
    progress = pyqtSignal(str)
    finished = pyqtSignal(bool, str)

    def __init__(self, meetings_data: list, force_download: bool, save_path: str, cache_dir: str, parent=None):
        super().__init__(parent)
        self.meetings_data = meetings_data
        self.force_download = force_download
        self.save_path = save_path
        self.cache_dir = Path(cache_dir)

    def run(self):
        try:
            all_dfs = []

            # Utilize the global NetworkSession so we inherit proxies and Humanness Delays
            session = NetworkSession.get_instance()
            NetworkSession.apply_humanness(session)

            total = len(self.meetings_data)

            for idx, mtg in enumerate(self.meetings_data):
                mtg_id = mtg.get("mtg_id")
                if not mtg_id:
                    continue  # Skip if we don't have a 3GPP Portal ID

                wg = mtg.get("wg_name", "")
                mtg_num = mtg.get("meeting_number", "")
                folder_name = mtg.get("folder_name") or mtg_num
                start_date = mtg.get("start_date", "")
                end_date = mtg.get("end_date", "")

                self.progress.emit(f"Processing {wg} {mtg_num} ({idx + 1}/{total})...")

                agenda_dir = self.cache_dir / folder_name / "Agenda"
                agenda_dir.mkdir(parents=True, exist_ok=True)

                # Find cached file
                filepath = self._find_cached_file(agenda_dir, mtg_id)

                # Check if we need to force download or if it's missing
                if self.force_download or not filepath or not filepath.exists():
                    self.progress.emit(f"Downloading {wg} {mtg_num} ({idx + 1}/{total})...")
                    filepath = self._download_sync(session, mtg_id, agenda_dir)

                if not filepath or not filepath.exists():
                    logging.warning(f"Skipping {wg} {mtg_num}: Could not resolve Excel file.")
                    continue

                # Parse the Excel file using your existing parser
                parsed_data = TDocsParser.parse_tdocs_excel(str(filepath))
                if not parsed_data:
                    continue

                # Convert to a DataFrame
                df = pd.DataFrame(parsed_data)

                # Inject the 4 new columns at the very front of the table
                df.insert(0, "WG", wg)
                df.insert(1, "Meeting", mtg_num)
                df.insert(2, "Start Date", start_date)
                df.insert(3, "End Date", end_date)

                all_dfs.append(df)

            if not all_dfs:
                self.finished.emit(False, "No TDocs found to merge for the selected meetings.")
                return

            self.progress.emit("Concatenating and saving master Excel file...")

            # Concat perfectly handles varying column layouts across different WG files
            master_df = pd.concat(all_dfs, ignore_index=True)

            # Save the raw data using pandas
            master_df.to_excel(self.save_path, index=False)

            self.progress.emit("Applying precise 3GPP TDoc formatting...")
            self._format_excel(self.save_path)

            self.finished.emit(True,
                               f"Successfully merged {len(master_df)} TDocs across {len(all_dfs)} meetings!\n\nSaved to:\n{self.save_path}")

        except Exception as e:
            logging.error(f"Error merging TDocs: {e}", exc_info=True)
            self.finished.emit(False, f"Error merging TDocs:\n{str(e)}")

    def _format_excel(self, filepath: str):
        """Applies formatting to match the official 3GPP TDocs list template."""
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        # 1. Exact 3GPP Styles extracted directly from the SA2 source file
        # Note: We use white text for the header to contrast the dark green background
        header_font = Font(name='Arial', size=9, bold=True, color="FFFFFFFF")
        header_fill = GradientFill(type="linear", stop=[Stop("FF75B91A", 0), Stop("FF54AF13", 1)])

        data_font = Font(name='Arial', size=8)
        data_alignment = Alignment(vertical='top', wrapText=True)

        # 2. Exact Column Widths mapped from the source file
        col_widths = {
            'WG': 10.0, 'Meeting': 14.0, 'Start Date': 12.0, 'End Date': 12.0,
            'TDoc': 9.14, 'Title': 36.57, 'Source': 14.0, 'Contact': 12.85,
            'Contact ID': 9.42, 'Type': 15.85, 'For': 15.85, 'Abstract': 16.85,
            'Secretary Remarks': 13.0, 'Agenda item sort order': 9.71,
            'Agenda item': 13.42, 'Agenda item description': 24.57,
            'TDoc sort order within agenda item': 16.0, 'TDoc Status': 20.14,
            'Reservation date': 16.14, 'Uploaded': 16.14, 'Is revision of': 14.28,
            'Revised to': 14.28, 'Release': 12.57, 'Spec': 9.85, 'Version': 13.0,
            'Related WIs': 16.42, 'CR': 10.42, 'CR revision': 13.0, 'CR category': 10.42,
            'TSG CR Pack': 13.0, 'UICC': 10.42, 'ME': 13.0, 'RAN': 13.0, 'CN': 13.0,
            'Clauses Affected': 13.0, 'Reply to': 15.42, 'To': 13.0, 'Cc': 13.0,
            'Original LS': 13.0, 'Reply in': 13.0
        }

        # 3. Format Headers and Set Widths
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = data_alignment  # Keep headers top-aligned and wrapped

            # Apply mapped width or default to 15.0 if column name is unknown
            header_val = str(cell.value).strip() if cell.value else ""
            width = col_widths.get(header_val, 15.0)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # 4. Format Data Rows
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = data_font
                cell.alignment = data_alignment

        # 5. UX Enhancements: Add Auto-Filter and Freeze the Top Row
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"

        wb.save(filepath)

    def _find_cached_file(self, agenda_dir: Path, mtg_id: str) -> Path:
        """Looks for the existing Excel file in the local cache."""
        if agenda_dir.exists():
            for f in agenda_dir.iterdir():
                if ("tdoc_list_meeting_" in f.name.lower() or "tdocs_list_" in f.name.lower()) and f.name.endswith(
                        ".xlsx"):
                    return f
        return agenda_dir / f"TDoc_List_Meeting_{mtg_id}.xlsx"

    def _download_sync(self, session, mtg_id: str, agenda_dir: Path):
        """Downloads the file synchronously within this background thread."""
        url = f"https://portal.3gpp.org/ngppapp/GenerateDocumentList.Aspx?meetingId={mtg_id}"
        try:
            response = session.get(url, stream=True, timeout=45)
            response.raise_for_status()

            filename = f"TDoc_List_Meeting_{mtg_id}.xlsx"
            content_disposition = response.headers.get('content-disposition')
            if content_disposition:
                matches = re.findall(r'filename="?([^"]+)"?', content_disposition)
                if matches:
                    filename = matches[0]

            filepath = agenda_dir / filename
            with open(filepath, 'wb') as f:
                for chunk in response.iter_content(chunk_size=65536):
                    if chunk:
                        f.write(chunk)
            return filepath
        except Exception as e:
            logging.error(f"Sync download failed for meeting {mtg_id}: {e}")
            return None