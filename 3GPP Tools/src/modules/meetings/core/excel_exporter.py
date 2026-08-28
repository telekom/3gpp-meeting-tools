# --- File: src/modules/meetings/core/excel_exporter.py ---
import os
import re
import logging
from pathlib import Path
from PyQt5.QtCore import QThread, pyqtSignal

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelExporterThread(QThread):
    progress = pyqtSignal(str)
    finished = pyqtSignal(bool, str)

    def __init__(
        self,
        output_path: Path,
        rows_data: list,
        selected_columns: list,
        mtg_info: dict,
        docs_ftp_url: str = "",
        auto_open: bool = True
    ):
        super().__init__()
        self.output_path = Path(output_path)
        self.rows_data = rows_data
        self.selected_columns = selected_columns
        self.mtg_info = mtg_info or {}
        self.docs_ftp_url = docs_ftp_url
        self.auto_open = auto_open

    def _normalize_docs_url(self) -> str:
        """Ensures the public 3GPP meeting documents URL is fully qualified."""
        url = self.docs_ftp_url or self.mtg_info.get("docs_folder_url", "")
        if url and not url.startswith("http"):
            url = "https://www.3gpp.org/ftp/" + url.lstrip('/')
        return url.rstrip('/')

    def run(self):
        try:
            self.progress.emit("Initializing Deutsche Telekom styled workbook...")
            wb = openpyxl.Workbook()
            ws = wb.active

            # --- Sheet Title Sanitization ---
            wg_name = str(self.mtg_info.get("wg_name", "3GPP")).strip()
            mtg_num = str(self.mtg_info.get("meeting_number", "")).strip()
            sheet_title = f"{wg_name} {mtg_num}".strip()[:31]
            ws.title = re.sub(r'[\\/*?:\[\]]', '_', sheet_title) or "TDocs"
            ws.views.sheetView[0].showGridLines = True

            # --- Corporate Styling Palette (#E20074 Magenta) ---
            magenta_hex = "E20074"
            magenta_dark_hex = "B8005E"
            zebra_tint_hex = "FDF5F8"  # Very soft magenta tint for alternating rows
            border_subtle_hex = "E8D0DC"

            header_fill = PatternFill(start_color=magenta_hex, end_color=magenta_hex, fill_type="solid")
            header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")

            zebra_fill = PatternFill(start_color=zebra_tint_hex, end_color=zebra_tint_hex, fill_type="solid")
            white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

            thin_border = Border(
                left=Side(style='thin', color=border_subtle_hex),
                right=Side(style='thin', color=border_subtle_hex),
                top=Side(style='thin', color=border_subtle_hex),
                bottom=Side(style='thin', color=border_subtle_hex)
            )
            header_border = Border(
                left=Side(style='thin', color=magenta_dark_hex),
                right=Side(style='thin', color=magenta_dark_hex),
                top=Side(style='thin', color=magenta_dark_hex),
                bottom=Side(style='medium', color=magenta_dark_hex)
            )

            # --- Alignments & Fonts ---
            center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

            regular_font = Font(name="Segoe UI", size=9)
            hyperlink_font = Font(name="Segoe UI", size=9, bold=True, color=magenta_hex, underline="single")

            center_cols = {"Type", "For", "Agenda Item", "TDoc Status", "My Status"}
            docs_base_url = self._normalize_docs_url()

            # --- 1. Write Header Row ---
            self.progress.emit("Writing headers...")
            ws.row_dimensions[1].height = 28

            for col_idx, col_name in enumerate(self.selected_columns, start=1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = header_border

            # --- 2. Write Data Rows ---
            total_rows = len(self.rows_data)
            for row_idx, row_dict in enumerate(self.rows_data, start=2):
                if row_idx % 250 == 0:
                    self.progress.emit(f"Writing row {row_idx - 1} of {total_rows}...")

                ws.row_dimensions[row_idx].height = 22
                is_zebra = (row_idx % 2 == 0)
                current_fill = zebra_fill if is_zebra else white_fill

                for col_idx, col_name in enumerate(self.selected_columns, start=1):
                    raw_val = self._extract_cell_value(col_name, row_dict)
                    cell = ws.cell(row=row_idx, column=col_idx, value=raw_val)
                    cell.fill = current_fill
                    cell.border = thin_border

                    # Hyperlink formatting for TDoc column
                    if col_name == "TDoc" and raw_val:
                        tdoc_id = str(raw_val).strip()
                        if docs_base_url:
                            cell.hyperlink = f"{docs_base_url}/{tdoc_id}.zip"
                        cell.font = hyperlink_font
                        cell.alignment = center_align
                    else:
                        cell.font = regular_font
                        cell.alignment = center_align if col_name in center_cols else left_align

            # --- 3. Optimized Column Widths ---
            self.progress.emit("Formatting column layout...")
            for col_idx, col_name in enumerate(self.selected_columns, start=1):
                col_letter = get_column_letter(col_idx)

                if col_name == "Source":
                    # Fixed compact width with text wrapping enabled
                    ws.column_dimensions[col_letter].width = 22
                elif col_name == "TDoc":
                    ws.column_dimensions[col_letter].width = 16
                elif col_name == "Title":
                    ws.column_dimensions[col_letter].width = 45
                elif col_name in ["Secretary Remarks", "My Notes", "Abstract"]:
                    ws.column_dimensions[col_letter].width = 40
                elif col_name == "Related TDocs":
                    ws.column_dimensions[col_letter].width = 24
                elif col_name in ["Type", "For", "Agenda Item", "TDoc Status", "My Status"]:
                    ws.column_dimensions[col_letter].width = 14
                else:
                    max_len = len(str(col_name))
                    for row_idx in range(2, min(total_rows + 2, 50)):
                        val = ws.cell(row=row_idx, column=col_idx).value
                        if val:
                            max_len = max(max_len, len(str(val).split('\n')[0]))
                    ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 30)

            # --- 4. Auto-Filter & Freeze Top Pane ---
            if total_rows > 0:
                ws.auto_filter.ref = ws.dimensions
            ws.freeze_panes = "A2"

            # --- 5. Save Spreadsheet ---
            self.progress.emit("Saving Excel file...")
            self.output_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(str(self.output_path))

            self.finished.emit(True, str(self.output_path))

        except PermissionError:
            self.finished.emit(
                False,
                f"Cannot write to '{self.output_path.name}'. Please close the spreadsheet if open in Excel."
            )
        except Exception as e:
            logging.error(f"[ExcelExporter] Error generating Excel: {e}", exc_info=True)
            self.finished.emit(False, str(e))

    def _extract_cell_value(self, col_name: str, row: dict) -> str:
        if col_name == "Related TDocs":
            parts = []
            if r_of := row.get("Is revision of"): parts.append(f"Rev of: {r_of}")
            if r_to := row.get("Revised to"): parts.append(f"Rev to: {r_to}")
            if r_ls := row.get("Original LS"): parts.append(f"Orig LS: {r_ls}")
            if r_re := row.get("Reply in"): parts.append(f"Reply: {r_re}")
            return "\n".join(parts)

        val = row.get(col_name, "")
        if val is None:
            return ""

        val_str = str(val).strip()
        if col_name == "My Status" and val_str == "⚪ Neutral":
            return ""

        return val_str